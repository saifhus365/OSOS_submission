import os
import datetime
import pandas as pd
import fitz  # PyMuPDF for PDFs
from docx import Document
from openpyxl import load_workbook

def get_pdf_metadata(filepath):
    try:
        doc = fitz.open(filepath)
        meta = doc.metadata or {}
        return {
            "title": meta.get("title"),
            "author": meta.get("author"),
            "created": meta.get("creationDate"),
            "modified": meta.get("modDate"),
            "producer": meta.get("producer"),
            "num_pages": doc.page_count,
        }
    except Exception as e:
        return {"error": f"PDF metadata extraction failed: {e}"}

def get_docx_metadata(filepath):
    try:
        doc = Document(filepath)
        core = doc.core_properties
        return {
            "title": core.title,
            "author": core.author,
            "created": core.created.strftime("%Y-%m-%d %H:%M:%S") if core.created else None,
            "last_modified_by": core.last_modified_by,
            "revision": core.revision,
            "subject": core.subject,
            "keywords": core.keywords,
        }
    except Exception as e:
        return {"error": f"DOCX metadata extraction failed: {e}"}

def get_csv_metadata(filepath):
    try:
        df = pd.read_csv(filepath, nrows=1000)  # limit to avoid memory issues
        return {
            "num_rows": df.shape[0],
            "num_columns": df.shape[1],
            "column_names": list(df.columns),
        }
    except Exception as e:
        return {"error": f"CSV metadata extraction failed: {e}"}

def get_excel_metadata(filepath):
    try:
        wb = load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        meta = wb.properties
        return {
            "sheet_names": sheets,
            "num_sheets": len(sheets),
            "creator": meta.creator,
            "created": meta.created.strftime("%Y-%m-%d %H:%M:%S") if meta.created else None,
            "last_modified_by": meta.lastModifiedBy,
        }
    except Exception as e:
        return {"error": f"Excel metadata extraction failed: {e}"}

def extract_extended_metadata(filepath):
    ext = os.path.splitext(filepath)[-1].lower()
    base_meta = {
        "source": os.path.basename(filepath),
        "extension": ext.lstrip("."),
    }

    if ext == ".pdf":
        base_meta.update(get_pdf_metadata(filepath))
    elif ext == ".docx":
        base_meta.update(get_docx_metadata(filepath))
    elif ext == ".csv":
        base_meta.update(get_csv_metadata(filepath))
    elif ext in [".xlsx", ".xls", ".xlsm"]:
        base_meta.update(get_excel_metadata(filepath))
    else:
        base_meta["warning"] = f"No metadata extractor available for {ext}"

    return base_meta
